from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from xlrd import open_workbook

list = []
list.append(648954584)
list.append(65533333)
list.append(694944)
list.append(1)
list.append(4)
list.append(2)
index=sorted(range(len(list)), key=lambda kk: list[kk])
print (index)

wb1 = open_workbook('1.xls')
table = wb1.sheets()[0]
sheet = wb1.sheet_names()[0]
lesson = table.cell(0,1).value
teacher = table.cell(0,4).value
nrows = table.nrows

wb = load_workbook(filename = '2.xlsx')
a = wb.get_sheet_names()
ws = wb[a[0]]
ws1 = wb.create_sheet(lesson)
ws1.sheet_properties.tabColor =colors.YELLOW
fontname=ws['B1'].font.name
fontsize=ws['B1'].font.size

ft_Red = Font(name=fontname,size=fontsize,color=colors.RED)
ft_Blue = Font(name=fontname,size=fontsize,color='FF00B0F0')
ft_Black = Font(name=fontname,size=fontsize,color=colors.BLACK)
rows = ws.max_row
columns = ws.max_column

for i in range(1,columns+1):
    ws1.cell(row=1, column=i).value=ws.cell(row=1, column=i).value
    ws1.cell(row=1, column=i).font=ft_Black
ws1['J1'].font = Font(name=fontname,size=fontsize, color=colors.RED,bold=True)
ws1['K1'].font = Font(name=fontname,size=fontsize,color='FF00B0F0',bold=True)
nw=1

for k in range(3,nrows-4):
    student = table.cell(k-1,1).value
    if type(student)== float or type(student)== int:
        student=int(student)
        student=str(student)
    test1= table.cell(k-1,4).value
    test2= table.cell(k-1,5).value
    name=table.cell(k-1,2).value
    nameJ=table.cell(k-1,3).value
    if len(student)==9:
        student='0'+student
    if len(student)==10:
        error=1
        for i in range(2,rows+1):
            if ws.cell(row=i, column=3).value == student and ws.cell(row=i, column=2).value == lesson and ws.cell(row=i, column=6).value == teacher:
                error=0
                nw=nw+1
                if test1 == 1:
                    ws1.cell(row=nw, column=columns-1).value='1.'+ws.cell(row=1, column=columns-1).value
                    ws1.cell(row=nw, column=columns-1).font=ft_Red
                    for j in range(1,columns-1):
                        ws1.cell(row=nw, column=j).value=ws.cell(row=i, column=j).value
                        ws1.cell(row=nw, column=j).font=ft_Red
                if test2 == 2:
                    ws1.cell(row=nw, column=columns).value='2.'+ws.cell(row=1, column=columns).value
                    ws1.cell(row=nw, column=columns).font=ft_Blue
                    for j in range(1,columns-1):
                        ws1.cell(row=nw, column=j).value=ws.cell(row=i, column=j).value
                        ws1.cell(row=nw, column=j).font=ft_Blue
                if name != ws1.cell(row=nw, column=4).value:
                    ws1.cell(row=nw, column=4).font=ft_Black
                if nameJ != ws1.cell(row=nw, column=5).value:
                    ws1.cell(row=nw, column=5).font=ft_Black
        if error==1:
                print ('Error in row=',k,'student number=',student)

wb.save(filename = '2.xlsx')